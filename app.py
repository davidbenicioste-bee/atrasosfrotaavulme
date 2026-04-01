import io
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st
import pandas as pd

# ──────────────────────────────────────────────────────────────
#  CONFIGURAÇÃO DA PÁGINA
# ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Extrator M.EMB.",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ──────────────────────────────────────────────────────────────
#  CSS CUSTOMIZADO - VERSÃO CLARA E ORGANIZADA
# ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Fontes modernas */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Fundo claro */
.stApp {
    background: linear-gradient(135deg, #f5f7fa 0%, #e9eef3 100%);
    color: #1e293b;
}

/* Esconde elementos padrão */
#MainMenu, footer, header { visibility: hidden; }

/* ── Cabeçalho moderno ── */
.hero {
    background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
    border: 1px solid #e2e8f0;
    border-radius: 24px;
    padding: 40px 48px;
    margin-bottom: 32px;
    box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05), 0 2px 4px -1px rgba(0,0,0,0.03);
}
.hero-badge {
    display: inline-block;
    background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
    color: white;
    font-family: 'Inter', monospace;
    font-size: 11px;
    font-weight: 600;
    letter-spacing: .05em;
    padding: 4px 14px;
    border-radius: 30px;
    margin-bottom: 16px;
}
.hero h1 {
    font-size: 32px;
    font-weight: 700;
    color: #0f172a;
    margin: 0 0 8px;
    letter-spacing: -0.02em;
}
.hero h1 span { 
    background: linear-gradient(135deg, #3b82f6, #06b6d4);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero p {
    color: #475569;
    font-size: 15px;
    font-weight: 400;
    margin: 0;
    line-height: 1.5;
}

/* ── Cards de seção ── */
.section-card {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 20px;
    padding: 28px 32px;
    margin-bottom: 24px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    transition: box-shadow 0.2s ease;
}
.section-card:hover {
    box-shadow: 0 4px 12px rgba(0,0,0,0.08);
}
.section-title {
    font-family: 'Inter', sans-serif;
    font-size: 13px;
    font-weight: 600;
    color: #3b82f6;
    text-transform: uppercase;
    letter-spacing: .08em;
    border-left: 3px solid #3b82f6;
    padding-left: 12px;
    margin-bottom: 24px;
}

/* ── Métricas modernas ── */
.metrics-row {
    display: flex;
    gap: 20px;
    margin-bottom: 28px;
    flex-wrap: wrap;
}
.metric-box {
    flex: 1;
    min-width: 160px;
    background: linear-gradient(135deg, #f8fafc 0%, #ffffff 100%);
    border: 1px solid #e2e8f0;
    border-radius: 16px;
    padding: 20px;
    text-align: center;
    transition: transform 0.2s ease;
}
.metric-box:hover {
    transform: translateY(-2px);
}
.metric-number {
    font-size: 38px;
    font-weight: 700;
    font-family: 'Inter', monospace;
    background: linear-gradient(135deg, #3b82f6, #06b6d4);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    line-height: 1;
    margin-bottom: 8px;
}
.metric-label {
    font-size: 12px;
    font-weight: 500;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: .05em;
}

/* ── Chips de núcleo ── */
.nucleo-chip {
    display: inline-block;
    background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
    border: 1px solid #bfdbfe;
    color: #1e40af;
    font-family: 'Inter', monospace;
    font-size: 12px;
    font-weight: 500;
    padding: 6px 14px;
    border-radius: 30px;
    margin: 4px 6px 4px 0;
    transition: all 0.2s ease;
}
.nucleo-chip:hover {
    background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%);
    transform: scale(1.02);
}

/* ── Tags de tipo ── */
.tag-m { 
    background: #fee2e2; 
    border: 1px solid #fecaca; 
    color: #991b1b; 
}
.tag-e { 
    background: #dcfce7; 
    border: 1px solid #bbf7d0; 
    color: #166534; 
}
.tag {
    display: inline-block;
    font-family: 'Inter', monospace;
    font-size: 11px;
    font-weight: 600;
    padding: 3px 12px;
    border-radius: 30px;
}

/* ── Caixas de info ── */
.info-box {
    background: #f1f5f9;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 14px 18px;
    color: #475569;
    font-size: 13px;
    margin-top: 8px;
}
.success-box {
    background: #f0fdf4;
    border: 1px solid #bbf7d0;
    border-radius: 12px;
    padding: 14px 18px;
    color: #166534;
    font-size: 13px;
    margin-bottom: 20px;
    font-weight: 500;
}
.warning-box {
    background: #fefce8;
    border: 1px solid #fde047;
    border-radius: 12px;
    padding: 14px 18px;
    color: #854d0e;
    font-size: 13px;
    margin-bottom: 12px;
}

/* ── Tabela de legenda ── */
.legend-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
    background: white;
    border-radius: 12px;
    overflow: hidden;
}
.legend-table th {
    background: #f8fafc;
    color: #1e293b;
    font-family: 'Inter', sans-serif;
    font-size: 12px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: .05em;
    padding: 12px 16px;
    text-align: left;
    border-bottom: 2px solid #e2e8f0;
}
.legend-table td {
    padding: 12px 16px;
    border-bottom: 1px solid #f1f5f9;
    color: #334155;
}
.legend-table tr:last-child td { border-bottom: none; }
.legend-table tr:hover td { background: #f8fafc; }

/* ── Botão de download ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
    border: none !important;
    color: white !important;
    font-weight: 600 !important;
    border-radius: 12px !important;
    padding: 12px 28px !important;
    font-size: 14px !important;
    width: 100% !important;
    transition: all 0.2s !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.05);
}
.stDownloadButton > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
}

/* ── File uploader ── */
.stFileUploader {
    background: #f8fafc;
    border: 2px dashed #cbd5e1;
    border-radius: 16px;
    padding: 8px;
}
[data-testid="stFileUploader"] {
    background: transparent;
}
.stFileUploader:hover {
    border-color: #3b82f6;
    background: #f1f5f9;
}

/* ── Streamlit dataframe ── */
.stDataFrame { 
    border-radius: 12px; 
    overflow: hidden;
    border: 1px solid #e2e8f0;
}

/* ── Separador ── */
hr { 
    border: none;
    border-top: 1px solid #e2e8f0;
    margin: 24px 0; 
}

/* ── Selectbox personalizado ── */
.stSelectbox label {
    color: #1e293b;
    font-weight: 500;
    font-size: 13px;
}
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────
#  FUNÇÕES DE EXTRAÇÃO
# ──────────────────────────────────────────────────────────────

def get_col_to_date(ws):
    col_date = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=5, column=c).value
        if isinstance(v, datetime.datetime):
            col_date[c] = v.strftime('%d/%m/%Y')
        elif isinstance(v, str) and v.isdigit():
            col_date[c] = f'{int(v):02d}/03/2026'
    return col_date


def extract_memb(wb_in):
    """
    Extrai registros M.EMB., anotações e linhas de legenda de todas as abas.
    Retorna: (records, legenda, nucleos_com_memb)
    """
    records = []
    legenda = []
    nucleos_com_memb = []

    LEGEND_KEYS = ['MAN - MANUTENÇÃO', 'MAN - EMBARCADA', 'AT OPE -', 'F-M-C -', 'TRÂNSITO -']

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        col_date = get_col_to_date(ws)

        # Identifica colunas de cabeçalho LINHA
        row5 = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
        linha_cols = [i + 1 for i, v in enumerate(row5) if v == 'LINHA']
        if not linha_cols:
            continue

        # Datas por período
        dates_p1 = {
            c: d for c, d in col_date.items()
            if c >= linha_cols[0] + 2 and c <= linha_cols[0] + 32
        }
        dates_p2 = {}
        if len(linha_cols) >= 2:
            dates_p2 = {
                c: d for c, d in col_date.items()
                if c >= linha_cols[1] + 2 and c <= linha_cols[1] + 32
            }

        # Mapa de comentários: (row, col) -> texto
        comments_map = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments_map[(cell.row, cell.column)] = cell.comment.text.strip()

        has_memb = False
        current_linha = None

        for row_num in range(6, ws.max_row + 1):
            col_a = ws.cell(row=row_num, column=1).value
            col_b = ws.cell(row=row_num, column=2).value

            # Atualiza linha atual
            if col_a and str(col_a).strip() not in ['TOTAL GERAL', 'Legenda', 'LINHA']:
                current_linha = str(col_a).strip()

            col_b_str = str(col_b).strip() if col_b else ''

            # ── Legenda ──
            if col_b_str and any(k in col_b_str.upper() for k in [l.upper() for l in LEGEND_KEYS]):
                tipo = 'Manutenção' if 'MANUTENÇÃO' in col_b_str.upper() else \
                       'Embarcada'  if 'EMBARCADA'  in col_b_str.upper() else 'Outro'
                entry = {'Núcleo': sheet_name, 'Descrição': col_b_str, 'Tipo': tipo}
                if entry not in legenda:
                    legenda.append(entry)

            # ── M.EMB. ──
            if 'M. EMB' in col_b_str.upper():
                has_memb = True

                for col, date_str in dates_p1.items():
                    val = ws.cell(row=row_num, column=col).value
                    if val and val != 0:
                        annot = comments_map.get((row_num, col), '')
                        records.append({
                            'Núcleo':      sheet_name,
                            'Linha':       current_linha,
                            'Período':     '1º Período (manhã)',
                            'Data':        date_str,
                            'Qtd Atrasos': val,
                            'Anotação':    annot.replace('\n', ' | '),
                        })

                for col, date_str in dates_p2.items():
                    val = ws.cell(row=row_num, column=col).value
                    if val and val != 0:
                        annot = comments_map.get((row_num, col), '')
                        records.append({
                            'Núcleo':      sheet_name,
                            'Linha':       current_linha,
                            'Período':     '2º Período (tarde)',
                            'Data':        date_str,
                            'Qtd Atrasos': val,
                            'Anotação':    annot.replace('\n', ' | '),
                        })

        if has_memb:
            nucleos_com_memb.append(sheet_name)

    return records, legenda, nucleos_com_memb


def build_excel(records, legenda):
    """Gera o arquivo Excel formatado com M.EMB. e Legenda."""
    wb = Workbook()

    # ── Aba principal: M.EMB. ──
    ws1 = wb.active
    ws1.title = 'M.EMB - Manut. Embarcada'

    h_fill  = PatternFill('solid', start_color='1F4E79')
    sh_fill = PatternFill('solid', start_color='2E75B6')
    a_fill  = PatternFill('solid', start_color='D6E4F0')
    w_fill  = PatternFill('solid', start_color='FFFFFF')
    thin    = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )

    # Título
    ws1.merge_cells('A1:F1')
    ws1['A1'].value = 'MAPA DE ATRASOS — MANUTENÇÃO EMBARCADA (M.EMB.)'
    ws1['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws1['A1'].fill = h_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    headers = ['Núcleo', 'Linha', 'Período', 'Data', 'Qtd Atrasos', 'Anotação']
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=2, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = sh_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
    ws1.row_dimensions[2].height = 20

    for i, rec in enumerate(records):
        ri   = i + 3
        fill = a_fill if i % 2 == 0 else w_fill
        vals = [rec['Núcleo'], rec['Linha'], rec['Período'],
                rec['Data'], rec['Qtd Atrasos'], rec['Anotação']]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name='Arial', size=10)
            cell.fill      = fill
            cell.border    = thin
            cell.alignment = Alignment(
                horizontal='center' if ci in [4, 5] else 'left',
                vertical='center',
                wrap_text=(ci == 6),
            )
        ws1.row_dimensions[ri].height = 30

    # Total
    tr = len(records) + 3
    ws1.merge_cells(f'A{tr}:D{tr}')
    ws1[f'A{tr}'].value     = 'TOTAL DE OCORRÊNCIAS'
    ws1[f'A{tr}'].font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    ws1[f'A{tr}'].fill      = h_fill
    ws1[f'A{tr}'].alignment = Alignment(horizontal='right', vertical='center')
    tv = ws1.cell(row=tr, column=5, value=sum(r['Qtd Atrasos'] for r in records))
    tv.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    tv.fill      = h_fill
    tv.alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=tr, column=6).fill = h_fill
    ws1.row_dimensions[tr].height = 22

    for col, w in zip('ABCDEF', [20, 14, 24, 14, 14, 72]):
        ws1.column_dimensions[col].width = w

    # ── Aba de Legenda ──
    if legenda:
        ws2 = wb.create_sheet('Legenda')

        ws2.merge_cells('A1:C1')
        ws2['A1'].value     = 'LEGENDA — MOTIVOS DE MANUTENÇÃO'
        ws2['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws2['A1'].fill      = h_fill
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 26

        for ci, h in enumerate(['Núcleo', 'Descrição', 'Tipo'], 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill      = sh_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = thin
        ws2.row_dimensions[2].height = 18

        for i, leg in enumerate(legenda):
            ri   = i + 3
            fill = a_fill if i % 2 == 0 else w_fill
            for ci, val in enumerate([leg['Núcleo'], leg['Descrição'], leg['Tipo']], 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name='Arial', size=10)
                cell.fill      = fill
                cell.border    = thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            ws2.row_dimensions[ri].height = 22

        for col, w in zip('ABC', [22, 48, 16]):
            ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────
#  INTERFACE STREAMLIT
# ──────────────────────────────────────────────────────────────

# Cabeçalho moderno
st.markdown("""
<div class="hero">
  <div class="hero-badge">⚙ FERRAMENTA INTERNA</div>
  <h1>Extrator <span>M.EMB.</span></h1>
  <p>Manutenção Embarcada — Extraia automaticamente todos os registros de atrasos<br>
  com datas, quantidades, anotações e legenda do Mapa de Atrasos.</p>
</div>
""", unsafe_allow_html=True)

# ── Upload ──
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📁 01 — Upload da Planilha</div>', unsafe_allow_html=True)

uploaded = st.file_uploader(
    label="",
    type=['xlsm', 'xlsx', 'xls'],
    help="Arraste o arquivo ou clique para selecionar. Aceita .xlsm e .xlsx até 50 MB.",
    label_visibility="collapsed",
)

st.markdown('<div class="info-box">📂 Aceita arquivos <b>.xlsm</b> e <b>.xlsx</b> — até 50 MB. A planilha deve seguir o formato padrão do Mapa de Atrasos com abas de Núcleo.</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ── Processamento ──
if uploaded is not None:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">⚙️ 02 — Processando</div>', unsafe_allow_html=True)

    with st.spinner('Lendo a planilha e extraindo dados M.EMB...'):
        try:
            raw = uploaded.read()
            wb  = openpyxl.load_workbook(io.BytesIO(raw), keep_vba=True, data_only=True)
            records, legenda, nucleos = extract_memb(wb)
        except Exception as e:
            st.error(f"❌ Erro ao ler o arquivo: {e}")
            st.stop()

    if not records:
        st.markdown('<div class="warning-box">⚠ Nenhum registro M.EMB. encontrado nesta planilha. Verifique se o arquivo segue o formato do Mapa de Atrasos.</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()

    total_atrasos = sum(r['Qtd Atrasos'] for r in records)

    st.markdown(f'<div class="success-box">✅ Extração concluída com sucesso! <b>{len(records)}</b> registros encontrados em <b>{len(nucleos)}</b> núcleo(s).</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Métricas ──
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">📊 03 — Resumo</div>', unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📋 Registros M.EMB.", len(records))
    with col2:
        st.metric("⏱ Total de Atrasos", total_atrasos)
    with col3:
        st.metric("🏢 Núcleos c/ M.EMB.", len(nucleos))
    with col4:
        st.metric("📝 Linhas únicas", len(set(r['Linha'] for r in records)))

    st.markdown("**Núcleos encontrados:**")
    chips_html = ''.join(f'<span class="nucleo-chip">{n}</span>' for n in nucleos)
    st.markdown(chips_html, unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Tabela de dados M.EMB. ──
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">📋 04 — Registros M.EMB.</div>', unsafe_allow_html=True)

    df = pd.DataFrame(records)

    # Filtros
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        nucleos_opcoes = ['Todos'] + sorted(df['Núcleo'].unique().tolist())
        filtro_nucleo = st.selectbox('Filtrar por Núcleo', nucleos_opcoes)
    with col_f2:
        linhas_opcoes = ['Todas'] + sorted(df['Linha'].unique().tolist())
        filtro_linha = st.selectbox('Filtrar por Linha', linhas_opcoes)
    with col_f3:
        periodos_opcoes = ['Todos'] + sorted(df['Período'].unique().tolist())
        filtro_periodo = st.selectbox('Filtrar por Período', periodos_opcoes)

    df_filtrado = df.copy()
    if filtro_nucleo  != 'Todos':  df_filtrado = df_filtrado[df_filtrado['Núcleo'] == filtro_nucleo]
    if filtro_linha   != 'Todas':  df_filtrado = df_filtrado[df_filtrado['Linha']  == filtro_linha]
    if filtro_periodo != 'Todos':  df_filtrado = df_filtrado[df_filtrado['Período'] == filtro_periodo]

    st.markdown(f"<small style='color:#64748b'>Exibindo <b>{len(df_filtrado)}</b> de <b>{len(df)}</b> registros</small>", unsafe_allow_html=True)

    st.dataframe(
        df_filtrado,
        use_container_width=True,
        hide_index=True,
        column_config={
            'Núcleo':      st.column_config.TextColumn('Núcleo',      width='medium'),
            'Linha':       st.column_config.TextColumn('Linha',       width='small'),
            'Período':     st.column_config.TextColumn('Período',     width='medium'),
            'Data':        st.column_config.TextColumn('Data',        width='small'),
            'Qtd Atrasos': st.column_config.NumberColumn('Qtd Atrasos', width='small', format="%d"),
            'Anotação':    st.column_config.TextColumn('Anotação',    width='large'),
        },
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Legenda ──
    if legenda:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">📌 05 — Legenda de Motivos</div>', unsafe_allow_html=True)

        rows_html = ''
        for leg in legenda:
            tipo_tag = (
                '<span class="tag tag-m">Manutenção</span>' if leg['Tipo'] == 'Manutenção'
                else '<span class="tag tag-e">Embarcada</span>' if leg['Tipo'] == 'Embarcada'
                else leg['Tipo']
            )
            rows_html += f"""
            <tr>
               <td>{leg['Núcleo']}</td>
               <td><code style="font-size:12px;color:#334155;background:#f1f5f9;padding:2px 6px;border-radius:4px">{leg['Descrição']}</code></td>
               <td>{tipo_tag}</td>
            </tr>"""

        st.markdown(f"""
        <table class="legend-table">
          <thead>
            <tr><th>Núcleo</th><th>Descrição</th><th>Tipo</th></tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Download ──
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">💾 06 — Exportar</div>', unsafe_allow_html=True)

    excel_buf = build_excel(records, legenda)
    nome_arquivo = f"MEMB_Manutencao_Embarcada.xlsx"

    st.download_button(
        label="⬇ Baixar Excel (.xlsx)",
        data=excel_buf,
        file_name=nome_arquivo,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
    )

    st.markdown(f"<div class='info-box' style='margin-top:12px'>📄 O arquivo Excel contém <b>2 abas</b>: <b>M.EMB - Manut. Embarcada</b> (com todos os {len(records)} registros) e <b>Legenda</b> (com os {len(legenda)} itens de legenda encontrados).</div>", unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)